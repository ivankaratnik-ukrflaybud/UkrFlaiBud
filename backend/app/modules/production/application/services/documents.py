from __future__ import annotations

# ruff: noqa: E501
from datetime import date, datetime
from decimal import Decimal
from html import escape
from io import BytesIO
from re import sub
from typing import Any
from uuid import UUID

from app.database.audit import AuditAction
from app.modules.production.application.services.common import ProductionServiceBase
from app.modules.production.infrastructure.repositories import (
    ProductionCompletionRepository,
    ProductionRequirementRepository,
    ProductionSnapshotRepository,
    ProductionStageRepository,
)

COMPANY_NAME = "ТОВ «Укрфлайбуд»"
DOCUMENT_TITLE = "ВИРОБНИЧЕ ЗАМОВЛЕННЯ"
WORKSHEET_NAME = "Замовлення"
PDF_FONT_NAME = "UkrFlyBudSans"


class ProductionDocumentService(ProductionServiceBase):
    async def prepare_document(self, order_id: UUID, *, user: Any | None = None) -> dict[str, Any]:
        order = await self.get_order(order_id)
        await self.ensure_order_scope(order, user)
        snapshot = await ProductionSnapshotRepository(self.session).get_for_order(order_id)
        requirements = await ProductionRequirementRepository(self.session).list_for_order(order_id)
        stages = await ProductionStageRepository(self.session).list_for_order(order_id)
        completions = await ProductionCompletionRepository(self.session).list_for_order(order_id)
        return {
            "order": {
                "id": str(order.id),
                "order_number": order.order_number,
                "name": order.name,
                "status": order.status,
                "priority": order.priority,
                "planned_quantity": str(order.planned_quantity),
                "completed_quantity": str(order.completed_quantity),
                "rejected_quantity": str(order.rejected_quantity),
                "planned_start_date": _iso(order.planned_start_date),
                "planned_end_date": _iso(order.planned_end_date),
                "notes": order.notes,
            },
            "snapshot": {
                "specification_code": snapshot.specification_code if snapshot else "",
                "specification_name": snapshot.specification_name if snapshot else "",
                "source_bom_version_number": snapshot.source_bom_version_number if snapshot else "",
                "product_code": snapshot.product_code if snapshot else "",
                "product_name": snapshot.product_name if snapshot else "",
                "unit_symbol": snapshot.unit_symbol if snapshot else "",
            },
            "requirements": [
                {
                    "line_number": item.line_number,
                    "item_code_snapshot": item.item_code_snapshot or "",
                    "display_name": item.display_name,
                    "planned_quantity": str(item.planned_quantity),
                    "reserved_quantity": str(item.reserved_quantity),
                    "issued_quantity": str(item.issued_quantity),
                    "returned_quantity": str(item.returned_quantity),
                    "consumed_quantity": str(item.consumed_quantity),
                    "scrapped_quantity": str(item.scrapped_quantity),
                    "unit_symbol_snapshot": item.unit_symbol_snapshot,
                    "source_type": item.source_type,
                }
                for item in requirements
            ],
            "stages": [
                {
                    "sequence": stage.sequence,
                    "name": stage.name,
                    "status": stage.status,
                    "progress_percent": stage.progress_percent,
                    "planned_start_at": _iso(stage.planned_start_at),
                    "planned_end_at": _iso(stage.planned_end_at),
                }
                for stage in stages
            ],
            "completions": [
                {
                    "completion_number": completion.completion_number,
                    "quantity_completed": str(completion.quantity_completed),
                    "quantity_rejected": str(completion.quantity_rejected),
                    "posted_at": _iso(completion.posted_at),
                }
                for completion in completions
            ],
        }

    async def audit_export(self, order_id: UUID, *, suffix: str, actor_id: UUID) -> None:
        await self.audit(
            action=(
                AuditAction.EXPORT_PDF.value if suffix == "pdf" else AuditAction.EXPORT_XLSX.value
            ),
            entity_type="production_order",
            entity_id=order_id,
            actor_id=actor_id,
            after={"suffix": suffix},
        )


def export_filename(order_number: str, suffix: str) -> str:
    safe_number = sub(r"[^A-Za-z0-9А-Яа-яІіЇїЄєҐґ_-]+", "_", order_number).strip("_") or "PO"
    return f"PO_{safe_number}_{date.today().isoformat()}.{suffix}"


def render_preview_html(document: dict[str, Any], *, include_toolbar: bool = True) -> str:
    order = document["order"]
    snapshot = document["snapshot"]
    toolbar = (
        """<div class="toolbar"><button onclick="window.print()">Друкувати</button>
<a href="./export/pdf">PDF</a><a href="./export/xlsx">Excel</a></div>"""
        if include_toolbar
        else ""
    )
    requirement_rows = "".join(
        "<tr>"
        f"<td>{item['line_number']}</td>"
        f"<td>{escape(item['item_code_snapshot'])}</td>"
        f"<td>{escape(item['display_name'])}</td>"
        f"<td>{_fmt(item['planned_quantity'])}</td>"
        f"<td>{_fmt(item['reserved_quantity'])}</td>"
        f"<td>{_fmt(item['issued_quantity'])}</td>"
        f"<td>{_fmt(item['returned_quantity'])}</td>"
        f"<td>{_fmt(item['consumed_quantity'])}</td>"
        f"<td>{_fmt(item['scrapped_quantity'])}</td>"
        f"<td>{escape(item['unit_symbol_snapshot'])}</td>"
        "</tr>"
        for item in document["requirements"]
    )
    stage_rows = "".join(
        "<tr>"
        f"<td>{stage['sequence']}</td><td>{escape(stage['name'])}</td>"
        f"<td>{escape(stage['status'])}</td><td>{stage['progress_percent']}%</td>"
        "</tr>"
        for stage in document["stages"]
    )
    completion_rows = "".join(
        "<tr>"
        f"<td>{item['completion_number']}</td><td>{_fmt(item['quantity_completed'])}</td>"
        f"<td>{_fmt(item['quantity_rejected'])}</td><td>{escape(item['posted_at'] or '')}</td>"
        "</tr>"
        for item in document["completions"]
    )
    return f"""<!doctype html>
<html lang="uk"><head><meta charset="utf-8" /><title>{escape(order["order_number"])}</title>
<style>
@page {{ size: A4 landscape; margin: 14mm; }}
body {{ font-family: "DejaVu Sans", Arial, sans-serif; color: #111827; margin: 0; }}
.toolbar {{ display: flex; gap: 8px; margin: 12px; }}
.toolbar a,.toolbar button {{ border: 1px solid #111827; background: white; color: #111827; padding: 8px 12px; text-decoration: none; }}
.document {{ margin: 0 auto; max-width: 210mm; padding: 14mm; }}
h1,h2,p {{ margin: 0; text-align: center; }}
h1 {{ font-size: 18px; margin-top: 4px; }}
h2 {{ font-size: 14px; margin: 5px 0 12px; }}
.meta {{ display: grid; grid-template-columns: 42mm 1fr 42mm 1fr; gap: 5px 10px; font-size: 11px; margin-bottom: 14px; }}
.meta strong {{ color: #374151; }}
table {{ border-collapse: collapse; font-size: 10px; margin-top: 10px; width: 100%; }}
thead {{ display: table-header-group; }}
tr {{ break-inside: avoid; }}
th,td {{ border: 1px solid #6b7280; padding: 4px; vertical-align: top; }}
th {{ background: #f3f4f6; text-align: center; }}
.signatures {{ display: grid; gap: 24px; grid-template-columns: 1fr 1fr 1fr; margin-top: 28px; }}
.line {{ border-bottom: 1px solid #111827; height: 28px; }}
@media print {{ .toolbar {{ display: none; }} .document {{ max-width: none; padding: 0; }} }}
</style></head><body>{toolbar}<section class="document production-order-print">
<p><strong>{COMPANY_NAME}</strong></p><h1>{DOCUMENT_TITLE}</h1><h2>{escape(order["order_number"])}</h2>
<div class="meta">
<strong>Виріб</strong><span>{escape(snapshot["product_name"])}</span>
<strong>Специфікація</strong><span>{escape(snapshot["specification_code"])} v{snapshot["source_bom_version_number"]}</span>
<strong>Планова кількість</strong><span>{_fmt(order["planned_quantity"])} {escape(snapshot["unit_symbol"])}</span>
<strong>Виконано</strong><span>{_fmt(order["completed_quantity"])}</span>
<strong>Пріоритет</strong><span>{escape(order["priority"])}</span>
<strong>Статус</strong><span>{escape(order["status"])}</span>
<strong>Початок</strong><span>{escape(order["planned_start_date"] or "")}</span>
<strong>Завершення</strong><span>{escape(order["planned_end_date"] or "")}</span>
</div>
<h3>Матеріальні потреби</h3><table><thead><tr><th>№</th><th>Код</th><th>Найменування</th><th>Потрібно</th><th>Резерв</th><th>Видано</th><th>Повернуто</th><th>Використано</th><th>Брак</th><th>Од.</th></tr></thead><tbody>{requirement_rows}</tbody></table>
<h3>Етапи</h3><table><thead><tr><th>№</th><th>Назва</th><th>Статус</th><th>Прогрес</th></tr></thead><tbody>{stage_rows}</tbody></table>
<h3>Оприбуткування</h3><table><thead><tr><th>№</th><th>Кількість</th><th>Брак</th><th>Дата</th></tr></thead><tbody>{completion_rows}</tbody></table>
<h3>Примітки</h3><p style="text-align:left">{escape(order["notes"] or "")}</p>
<div class="signatures"><div>Підготував<div class="line"></div></div><div>Виконав<div class="line"></div></div><div>Прийняв<div class="line"></div></div></div>
</section></body></html>"""


def render_xlsx(document: dict[str, Any]) -> bytes:
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import (  # type: ignore[import-untyped]
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = WORKSHEET_NAME
    order = document["order"]
    snapshot = document["snapshot"]
    sheet.merge_cells("A1:J1")
    sheet.merge_cells("A2:J2")
    sheet["A1"] = COMPANY_NAME
    sheet["A2"] = DOCUMENT_TITLE
    sheet["A4"] = "Номер"
    sheet["B4"] = order["order_number"]
    sheet["D4"] = "Виріб"
    sheet["E4"] = snapshot["product_name"]
    sheet["A5"] = "Специфікація"
    sheet["B5"] = f"{snapshot['specification_code']} v{snapshot['source_bom_version_number']}"
    sheet["D5"] = "План"
    sheet["E5"] = float(Decimal(order["planned_quantity"]))
    headers = [
        "№",
        "Код",
        "Найменування",
        "Потрібно",
        "Резерв",
        "Видано",
        "Повернуто",
        "Використано",
        "Брак",
        "Од.",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=8, column=column, value=header)
    row = 9
    for item in document["requirements"]:
        values = [
            item["line_number"],
            item["item_code_snapshot"],
            item["display_name"],
            float(Decimal(item["planned_quantity"])),
            float(Decimal(item["reserved_quantity"])),
            float(Decimal(item["issued_quantity"])),
            float(Decimal(item["returned_quantity"])),
            float(Decimal(item["consumed_quantity"])),
            float(Decimal(item["scrapped_quantity"])),
            item["unit_symbol_snapshot"],
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
        row += 1
    thin = Side(style="thin", color="6B7280")
    for row_cells in sheet.iter_rows(min_row=1, max_row=row - 1, min_col=1, max_col=10):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="DejaVu Sans", size=10)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[8]:
        cell.font = Font(name="DejaVu Sans", bold=True)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
    for row_number in (1, 2):
        sheet[f"A{row_number}"].font = Font(name="DejaVu Sans", bold=True, size=14)
        sheet[f"A{row_number}"].alignment = Alignment(horizontal="center")
    widths = [7, 18, 42, 14, 14, 14, 14, 14, 12, 10]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A9"
    sheet.print_title_rows = "8:8"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_area = f"A1:J{row - 1}"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def render_pdf(document: dict[str, Any]) -> bytes:
    from reportlab.lib import colors  # type: ignore[import-untyped]
    from reportlab.lib.pagesizes import A4, landscape  # type: ignore[import-untyped]
    from reportlab.lib.styles import getSampleStyleSheet  # type: ignore[import-untyped]
    from reportlab.lib.units import mm  # type: ignore[import-untyped]
    from reportlab.pdfbase import pdfmetrics  # type: ignore[import-untyped]
    from reportlab.pdfbase.ttfonts import TTFont  # type: ignore[import-untyped]
    from reportlab.platypus import (  # type: ignore[import-untyped]
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    from app.modules.bom.application.documents import _register_unicode_font

    font_name, _, _ = _register_unicode_font(pdfmetrics, TTFont)
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm)
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = font_name
    story: list[Any] = [
        Paragraph(COMPANY_NAME, styles["Title"]),
        Paragraph(DOCUMENT_TITLE, styles["Heading1"]),
        Paragraph(document["order"]["order_number"], styles["Heading2"]),
        Spacer(1, 8),
    ]
    rows = [["№", "Код", "Найменування", "Потрібно", "Видано", "Од."]]
    for item in document["requirements"]:
        rows.append(
            [
                item["line_number"],
                item["item_code_snapshot"],
                Paragraph(escape(item["display_name"]), styles["BodyText"]),
                _fmt(item["planned_quantity"]),
                _fmt(item["issued_quantity"]),
                item["unit_symbol_snapshot"],
            ]
        )
    table = Table(
        rows, repeatRows=1, colWidths=[10 * mm, 25 * mm, 75 * mm, 25 * mm, 25 * mm, 15 * mm]
    )
    table.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), font_name),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#6B7280")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.extend(
        [
            table,
            Spacer(1, 18),
            Paragraph("Підпис відповідального: ____________________", styles["BodyText"]),
        ]
    )
    pdf.build(story)
    return buffer.getvalue()


def _iso(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def _fmt(value: object) -> str:
    return format(Decimal(str(value)).normalize(), "f")
