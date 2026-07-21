from __future__ import annotations

# ruff: noqa: E501
from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal
from html import escape
from io import BytesIO
from pathlib import Path
from re import sub
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from app.modules.bom.domain.entities import BomVersionStatus

COMPANY_NAME = "ТОВ «Укрфлайбуд»"
DOCUMENT_TITLE = "СПЕЦИФІКАЦІЯ ВИРОБУ"
WORKSHEET_NAME = "Специфікація"
PDF_FONT_NAME = "UkrFlyBudSans"
PDF_FONT_BOLD_NAME = "UkrFlyBudSansBold"

SPEC_COLUMNS = [
    "№",
    "Код позиції",
    "Найменування",
    "Кількість",
    "Одиниця виміру",
    "Виробник / модель",
    "Примітка",
]

IMPORT_COLUMNS = [
    "Код позиції",
    "Найменування",
    "Кількість",
    "Одиниця виміру",
    "Виробник",
    "Артикул виробника",
    "Креслення",
    "Технічні вимоги",
    "Примітка",
]


def export_filename(code: str, version_number: int, suffix: str) -> str:
    safe_code = sub(r"[^A-Za-z0-9А-Яа-яІіЇїЄєҐґ_-]+", "_", code).strip("_") or "SPEC"
    return f"SPEC_{safe_code}_v{version_number}_{date.today().isoformat()}.{suffix}"


def render_preview_html(document: dict[str, Any], *, include_toolbar: bool = True) -> str:
    status_label = _status_label(document["version"]["status"])
    generated_at = _generated_at()
    watermark = (
        "draft" if document["version"]["status"] == BomVersionStatus.DRAFT.value else "approved"
    )
    rows = "\n".join(_html_line(line) for line in document["lines"])
    toolbar = (
        """  <div class="toolbar">
    <button onclick="window.print()">Друкувати</button>
    <a href="../export/pdf">Завантажити PDF</a>
    <a href="../export/xlsx">Завантажити Excel</a>
  </div>
"""
        if include_toolbar
        else ""
    )
    return f"""<!doctype html>
<html lang="uk">
<head>
  <meta charset="utf-8" />
  <title>{escape(document["specification"]["code"])} v{document["version"]["version_number"]}</title>
  <style>
    @page {{ size: A4 portrait; margin: 14mm; }}
    * {{ box-sizing: border-box; }}
    html, body {{ margin: 0; padding: 0; }}
    body {{ font-family: "DejaVu Sans", Arial, sans-serif; color: #111827; background: #fff; }}
    .toolbar {{ display: flex; gap: 8px; margin-bottom: 18px; }}
    .toolbar a, .toolbar button {{ border: 1px solid #111827; background: #fff; color: #111827; cursor: pointer; padding: 8px 12px; text-decoration: none; }}
    .document {{ background: #fff; margin: 0 auto; max-width: 210mm; padding: 14mm; position: relative; }}
    .document-header {{ break-after: avoid; page-break-after: avoid; text-align: center; }}
    .company {{ font-size: 14px; font-weight: 700; margin: 0 0 5px; text-transform: uppercase; }}
    h1 {{ font-size: 18px; margin: 0 0 6px; }}
    h2 {{ font-size: 15px; font-weight: 700; margin: 0 0 14px; }}
    .status {{ border: 1px solid #111827; display: inline-block; font-size: 11px; font-weight: 700; margin-bottom: 10px; padding: 3px 8px; text-transform: uppercase; }}
    .mark {{ font-size: 64px; left: 16%; opacity: .06; position: fixed; text-transform: uppercase; top: 42%; transform: rotate(-28deg); z-index: 0; }}
    .meta {{ break-after: avoid; display: grid; font-size: 11px; gap: 5px 10px; grid-template-columns: 42mm 1fr 42mm 1fr; margin-bottom: 16px; position: relative; z-index: 1; }}
    .meta strong {{ color: #374151; }}
    table {{ border-collapse: collapse; font-size: 11px; width: 100%; }}
    thead {{ display: table-header-group; }}
    tbody {{ display: table-row-group; }}
    tr {{ break-inside: avoid; page-break-inside: avoid; }}
    th, td {{ border: 1px solid #6b7280; overflow-wrap: break-word; padding: 5px; vertical-align: top; }}
    th {{ background: #f3f4f6; font-weight: 700; text-align: center; }}
    td:nth-child(1), td:nth-child(4), td:nth-child(5) {{ text-align: center; }}
    .signatures {{ break-inside: avoid; display: grid; gap: 28px; grid-template-columns: 1fr 1fr 1fr; margin-top: 28px; page-break-inside: avoid; }}
    .line {{ border-bottom: 1px solid #111827; height: 28px; }}
    .footer {{ border-top: 1px solid #d1d5db; color: #374151; display: flex; font-size: 10px; justify-content: space-between; margin-top: 18px; padding-top: 5px; }}
    @media print {{
      .toolbar {{ display: none; }}
      body {{ background: #fff; color: #000; margin: 0; }}
      .document {{ box-shadow: none; margin: 0; max-width: none; padding: 0; }}
      a[href]::after {{ content: ""; }}
    }}
  </style>
</head>
<body>
{toolbar}
  <section class="document bom-print-document">
    <div class="mark {watermark}">{escape(status_label)}</div>
    <header class="document-header">
      <p class="company">{COMPANY_NAME}</p>
      <h1>{DOCUMENT_TITLE}</h1>
      <h2>{escape(document["specification"]["name"])}</h2>
      <div class="status">{escape(status_label)}</div>
    </header>
    <div class="meta">
      <strong>Код специфікації</strong><span>{escape(document["specification"]["code"])}</span>
      <strong>Виріб</strong><span>{escape(str(document.get("product_name") or ""))}</span>
      <strong>Версія</strong><span>{document["version"]["version_number"]}</span>
      <strong>Організація</strong><span>{COMPANY_NAME}</span>
      <strong>Статус</strong><span>{escape(status_label)}</span>
      <strong>Дата формування документа</strong><span>{escape(generated_at)}</span>
    </div>
    <table>
      <thead><tr>{"".join(f"<th>{escape(column)}</th>" for column in SPEC_COLUMNS)}</tr></thead>
      <tbody>{rows}</tbody>
    </table>
    <div class="signatures">
      <div><strong>Підготував</strong><div class="line"></div></div>
      <div><strong>Перевірив</strong><div class="line"></div></div>
      <div><strong>Затвердив</strong><div class="line"></div></div>
    </div>
    <footer class="footer"><span>{escape(document["specification"]["code"])} v{document["version"]["version_number"]}</span><span>{escape(generated_at)}</span></footer>
  </section>
</body>
</html>"""


def render_pdf(document: dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors  # type: ignore[import-untyped]
        from reportlab.lib.enums import TA_CENTER  # type: ignore[import-untyped]
        from reportlab.lib.pagesizes import A4, landscape, portrait  # type: ignore[import-untyped]
        from reportlab.lib.styles import (  # type: ignore[import-untyped]
            ParagraphStyle,
            getSampleStyleSheet,
        )
        from reportlab.lib.units import mm  # type: ignore[import-untyped]
        from reportlab.pdfbase import pdfmetrics  # type: ignore[import-untyped]
        from reportlab.pdfbase.ttfonts import TTFont  # type: ignore[import-untyped]
        from reportlab.platypus import (  # type: ignore[import-untyped]
            KeepTogether,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:  # pragma: no cover - dependency is installed in Docker.
        raise RuntimeError("PDF generation requires reportlab.") from exc

    font_name, bold_font_name, _ = _register_unicode_font(pdfmetrics, TTFont)
    status_label = _status_label(document["version"]["status"])
    generated_at = _generated_at()
    page_size = landscape(A4) if _table_requires_landscape(document) else portrait(A4)
    buffer = BytesIO()
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=16 * mm,
        title=f"{document['specification']['code']} v{document['version']['version_number']}",
    )
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            "BomCompany",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontName=bold_font_name,
            fontSize=11,
            leading=14,
        )
    )
    styles.add(
        ParagraphStyle(
            "BomTitle",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontName=bold_font_name,
            fontSize=14,
            leading=18,
        )
    )
    styles.add(
        ParagraphStyle(
            "BomSubtitle",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontName=bold_font_name,
            fontSize=11,
            leading=15,
        )
    )
    body = ParagraphStyle(
        "BomBody",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=8,
        leading=10,
    )
    label = ParagraphStyle(
        "BomLabel",
        parent=body,
        fontName=bold_font_name,
        textColor=colors.HexColor("#374151"),
    )
    header = KeepTogether(
        [
            Paragraph(COMPANY_NAME, styles["BomCompany"]),
            Paragraph(DOCUMENT_TITLE, styles["BomTitle"]),
            Paragraph(str(document["specification"]["name"]), styles["BomSubtitle"]),
            Spacer(1, 4),
            _pdf_status_table(status_label, bold_font_name, colors, Table, TableStyle),
            Spacer(1, 8),
        ]
    )
    metadata = Table(
        [
            [
                Paragraph("Код специфікації", label),
                Paragraph(str(document["specification"]["code"]), body),
                Paragraph("Виріб", label),
                Paragraph(str(document.get("product_name") or ""), body),
            ],
            [
                Paragraph("Версія", label),
                Paragraph(str(document["version"]["version_number"]), body),
                Paragraph("Організація", label),
                Paragraph(COMPANY_NAME, body),
            ],
            [
                Paragraph("Статус", label),
                Paragraph(status_label, body),
                Paragraph("Дата формування документа", label),
                Paragraph(generated_at, body),
            ],
        ],
        colWidths=[34 * mm, 54 * mm, 42 * mm, 58 * mm],
        hAlign="LEFT",
    )
    metadata.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), font_name),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F9FAFB")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F9FAFB")),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    table = _pdf_lines_table(
        document, body, font_name, bold_font_name, colors, mm, Table, TableStyle
    )
    signatures = Table(
        [
            [
                Paragraph("Підготував", label),
                Paragraph("Перевірив", label),
                Paragraph("Затвердив", label),
            ],
            ["", "", ""],
        ],
        colWidths=[55 * mm, 55 * mm, 55 * mm],
    )
    signatures.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), font_name),
                ("LINEBELOW", (0, 1), (-1, 1), 0.5, colors.black),
                ("TOPPADDING", (0, 1), (-1, 1), 18),
            ]
        )
    )
    story: list[Any] = [header, metadata, Spacer(1, 8), table, Spacer(1, 18), signatures]
    pdf.build(
        story, onFirstPage=_footer(font_name, document), onLaterPages=_footer(font_name, document)
    )
    return buffer.getvalue()


def render_xlsx(document: dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
        from openpyxl.styles import (  # type: ignore[import-untyped]
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover - dependency is installed in Docker.
        raise RuntimeError("XLSX generation requires openpyxl.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = WORKSHEET_NAME
    status_label = _status_label(document["version"]["status"])
    generated_at = _generated_at()
    sheet.merge_cells("A1:G1")
    sheet.merge_cells("A2:G2")
    sheet.merge_cells("A3:G3")
    sheet["A1"] = COMPANY_NAME
    sheet["A2"] = DOCUMENT_TITLE
    sheet["A3"] = document["specification"]["name"]
    sheet["A5"] = "Код"
    sheet["B5"] = document["specification"]["code"]
    sheet["A6"] = "Версія"
    sheet["B6"] = document["version"]["version_number"]
    sheet["A7"] = "Статус"
    sheet["B7"] = status_label
    sheet["D5"] = "Виріб"
    sheet["E5"] = document.get("product_name") or ""
    sheet.merge_cells("E5:G5")
    sheet["D6"] = "Організація"
    sheet["E6"] = COMPANY_NAME
    sheet.merge_cells("E6:G6")
    sheet["D7"] = "Дата"
    sheet["E7"] = generated_at
    sheet.merge_cells("E7:G7")
    header_row = 10
    for column_index, column in enumerate(SPEC_COLUMNS, start=1):
        sheet.cell(row=header_row, column=column_index, value=column)
    row_index = header_row + 1
    for line in document["lines"]:
        values = _line_values(line)
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
        row_index += 1
        if line.get("technical_requirements"):
            sheet.cell(
                row=row_index,
                column=3,
                value=f"Технічні вимоги: {line['technical_requirements']}",
            )
            row_index += 1

    thin = Side(style="thin", color="6B7280")
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    label_fill = PatternFill("solid", fgColor="F3F4F6")
    widths = [7, 18, 46, 14, 18, 30, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(1, row_index):
        for cell in sheet[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=row >= header_row)
            cell.font = Font(name="DejaVu Sans", size=10)
    for row in (1, 2, 3):
        sheet[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        sheet[f"A{row}"].font = Font(name="DejaVu Sans", bold=True, size=14 if row == 2 else 12)
        sheet.row_dimensions[row].height = 24 if row == 2 else 21
    sheet["A3"].font = Font(name="DejaVu Sans", bold=True, size=11)
    for cell_reference in ("A5", "A6", "A7", "D5", "D6", "D7"):
        sheet[cell_reference].font = Font(name="DejaVu Sans", bold=True)
        sheet[cell_reference].fill = label_fill
        sheet[cell_reference].alignment = Alignment(vertical="top", wrap_text=False)
    for row in sheet.iter_rows(min_row=5, max_row=7, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[header_row]:
        cell.font = Font(name="DejaVu Sans", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=row_index - 1, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:G{row_index - 1}"
    sheet.print_title_rows = f"{header_row}:{header_row}"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5
    sheet.print_area = f"A1:G{row_index - 1}"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def render_import_template() -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return _minimal_xlsx("Імпорт", IMPORT_COLUMNS)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Імпорт"
    sheet.append(IMPORT_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
    widths = [18, 42, 14, 18, 24, 24, 18, 42, 36]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + column)].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_import_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is installed in Docker.
        raise RuntimeError("XLSX import requires openpyxl.") from exc

    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    headers = [
        str(value).strip() if value is not None else ""
        for value in next(sheet.iter_rows(values_only=True))
    ]
    rows: list[dict[str, Any]] = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue
        row = {
            headers[position]: value
            for position, value in enumerate(values)
            if position < len(headers)
        }
        row["_row_number"] = index
        rows.append(row)
    return rows


def format_decimal(value: object) -> str:
    decimal_value = Decimal(str(value))
    return format(decimal_value.normalize(), "f")


def _html_line(line: dict[str, Any]) -> str:
    return (
        "<tr>"
        f"<td>{escape(str(line['line_number']))}</td>"
        f"<td>{escape(str(line.get('position_code') or ''))}</td>"
        f"<td>{escape(str(line['display_name']))}</td>"
        f"<td>{escape(format_decimal(line['quantity']))}</td>"
        f"<td>{escape(str(line.get('unit_symbol') or line.get('unit_name') or ''))}</td>"
        f"<td>{escape(str(line.get('manufacturer_summary') or ''))}</td>"
        f"<td>{escape(str(line.get('notes') or ''))}</td>"
        "</tr>"
    )


def _line_values(line: dict[str, Any]) -> list[object]:
    return [
        line["line_number"],
        line.get("position_code") or "",
        line["display_name"],
        float(Decimal(str(line["quantity"]))),
        line.get("unit_symbol") or line.get("unit_name") or "",
        line.get("manufacturer_summary") or "",
        line.get("notes") or "",
    ]


def _status_label(status: str) -> str:
    labels = {
        "draft": "Чернетка",
        "under_review": "На розгляді",
        "approved": "Затверджено",
        "superseded": "Замінено",
        "archived": "Архів",
    }
    return labels.get(status, status)


def _generated_at() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def _table_requires_landscape(document: dict[str, Any]) -> bool:
    return any(len(str(line.get("display_name") or "")) > 42 for line in document["lines"])


def _register_unicode_font(pdfmetrics: Any, ttfont: Any) -> tuple[str, str, Path]:
    candidates = [
        (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ),
        (
            Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
            Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
        ),
        (Path("C:/Windows/Fonts/arial.ttf"), Path("C:/Windows/Fonts/arialbd.ttf")),
    ]
    for regular_path, bold_path in candidates:
        if regular_path.exists():
            pdfmetrics.registerFont(ttfont(PDF_FONT_NAME, str(regular_path)))
            resolved_bold_path = bold_path if bold_path.exists() else regular_path
            pdfmetrics.registerFont(ttfont(PDF_FONT_BOLD_NAME, str(resolved_bold_path)))
            return PDF_FONT_NAME, PDF_FONT_BOLD_NAME, regular_path
    raise RuntimeError("No Unicode TrueType font is available for BOM PDF generation.")


def _pdf_status_table(
    status_label: str, bold_font_name: str, colors: Any, table_class: Any, table_style_class: Any
) -> Any:
    status_table = table_class([[status_label]], hAlign="CENTER")
    status_table.setStyle(
        table_style_class(
            [
                ("FONT", (0, 0), (-1, -1), bold_font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return status_table


def _pdf_lines_table(
    document: dict[str, Any],
    body_style: Any,
    font_name: str,
    bold_font_name: str,
    colors: Any,
    mm: Any,
    table_class: Any,
    table_style_class: Any,
) -> Any:
    data: list[list[Any]] = [[_paragraph(column, body_style) for column in SPEC_COLUMNS]]
    for line in document["lines"]:
        data.append([_paragraph(value, body_style) for value in _line_values(line)])
        if line.get("technical_requirements"):
            data.append(
                [
                    "",
                    "",
                    _paragraph(f"Технічні вимоги: {line['technical_requirements']}", body_style),
                    "",
                    "",
                    "",
                    "",
                ]
            )
    table = table_class(
        data,
        repeatRows=1,
        colWidths=[10 * mm, 24 * mm, 70 * mm, 22 * mm, 24 * mm, 46 * mm, 50 * mm],
        hAlign="LEFT",
    )
    table.setStyle(
        table_style_class(
            [
                ("FONT", (0, 0), (-1, -1), font_name),
                ("FONT", (0, 0), (-1, 0), bold_font_name),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#6B7280")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (3, 0), (4, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _paragraph(value: object, style: Any) -> Any:
    from reportlab.platypus import Paragraph

    return Paragraph(escape(str(value)), style)


def _footer(font_name: str, document: dict[str, Any]) -> Any:
    def draw(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setFont(font_name, 8)
        footer_text = (
            f"{document['specification']['code']} v{document['version']['version_number']}"
        )
        canvas.drawString(doc.leftMargin, 8, footer_text)
        canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 8, f"Сторінка {doc.page}")
        canvas.restoreState()

    return draw


def _minimal_xlsx(sheet_name: str, headers: Iterable[str]) -> bytes:
    shared_strings = list(headers)
    sheet_cells = "".join(
        f'<c r="{chr(65 + index)}1" t="s"><v>{index}</v></c>'
        for index, _ in enumerate(shared_strings)
    )
    files = {
        "[Content_Types].xml": """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
</Types>""",
        "_rels/.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""",
        "xl/workbook.xml": f"""<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="{escape(sheet_name)}" sheetId="1" r:id="rId1"/></sheets></workbook>""",
        "xl/_rels/workbook.xml.rels": """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
</Relationships>""",
        "xl/worksheets/sheet1.xml": f"""<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData><row r="1">{sheet_cells}</row></sheetData></worksheet>""",
        "xl/sharedStrings.xml": f"""<?xml version="1.0" encoding="UTF-8"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(shared_strings)}" uniqueCount="{len(shared_strings)}">
{"".join(f"<si><t>{escape(text)}</t></si>" for text in shared_strings)}</sst>""",
    }
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content.encode("utf-8"))
    return output.getvalue()
