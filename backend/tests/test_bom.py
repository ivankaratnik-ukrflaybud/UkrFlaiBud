from collections.abc import AsyncGenerator
from contextlib import asynccontextmanager
from io import BytesIO
from uuid import uuid4

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_unit_of_work
from app.core.config import settings
from app.database.audit import AuditLog
from app.main import create_app
from app.models.base import ConflictError, PermissionDeniedError, ValidationError
from app.modules.bom.application.documents import (
    WORKSHEET_NAME,
    export_filename,
    render_import_template,
    render_pdf,
    render_preview_html,
    render_xlsx,
)
from app.modules.bom.application.services import BomService
from app.modules.identity.application.security import create_access_token, decode_access_token
from app.modules.identity.application.services import (
    AuthService,
    IdentityService,
    require_permission_for_user,
)
from app.modules.identity.infrastructure.models import (
    PermissionModel,
    RoleModel,
    RolePermissionModel,
    UserModel,
)
from app.modules.inventory.application.services import CatalogService
from app.modules.organizations.application.services import OrganizationService
from app.repositories.sqlalchemy_unit_of_work import SQLAlchemyUnitOfWork
from app.schemas.pagination import PageRequest, SortDirection


class UnitOfWorkStub:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self._session = session

    async def commit(self) -> None:
        await self.session.commit()

    async def rollback(self) -> None:
        await self.session.rollback()

    async def flush(self) -> None:
        await self.session.flush()

    async def refresh(self, entity: object, attribute_names: list[str] | None = None) -> None:
        await self.session.refresh(entity, attribute_names=attribute_names)


@asynccontextmanager
async def service_context(db_session: AsyncSession) -> AsyncGenerator[SQLAlchemyUnitOfWork]:
    unit_of_work = UnitOfWorkStub(db_session)
    try:
        yield unit_of_work
    except Exception:
        await unit_of_work.rollback()
        raise


async def create_bom_fixture(db_session: AsyncSession) -> dict[str, object]:
    async with service_context(db_session) as unit_of_work:
        organization = await OrganizationService(unit_of_work).create(
            {
                "name": "UKRFLYBUD",
                "short_name": "UFB",
                "legal_name": "UKRFLYBUD LLC",
                "edrpou": "12345678",
                "is_active": True,
            }
        )
        other_organization = await OrganizationService(unit_of_work).create(
            {
                "name": "Other",
                "short_name": "OTH",
                "legal_name": "Other LLC",
                "edrpou": "87654321",
                "is_active": True,
            }
        )
        admin = UserModel(
            email="admin@example.com",
            normalized_email="admin@example.com",
            password_hash="hash",
            display_name="Admin",
            is_active=True,
            is_superuser=True,
        )
        viewer = UserModel(
            email="viewer@example.com",
            normalized_email="viewer@example.com",
            password_hash="hash",
            display_name="Viewer",
            is_active=True,
            is_superuser=False,
        )
        db_session.add_all([admin, viewer])
        await db_session.flush()
        unit = await CatalogService(unit_of_work).create_unit(
            {
                "organization_id": organization.id,
                "code": "PCS",
                "name": "штука",
                "symbol": "шт",
                "precision": 0,
            },
            actor_id=admin.id,
        )
        other_unit = await CatalogService(unit_of_work).create_unit(
            {
                "organization_id": other_organization.id,
                "code": "PCS",
                "name": "piece",
                "symbol": "pcs",
                "precision": 0,
            },
            actor_id=admin.id,
        )
        category = await CatalogService(unit_of_work).create_category(
            {"organization_id": organization.id, "code": "ELEC", "name": "Електроніка"},
            actor_id=admin.id,
        )
        other_category = await CatalogService(unit_of_work).create_category(
            {
                "organization_id": other_organization.id,
                "code": "ELEC",
                "name": "Electronics",
            },
            actor_id=admin.id,
        )
        item = await CatalogService(unit_of_work).create_item(
            {
                "organization_id": organization.id,
                "sku": "MOTOR-001",
                "name": "Двигун",
                "category_id": category.id,
                "unit_of_measure_id": unit.id,
                "item_type": "component",
                "manufacturer": "UFB",
                "manufacturer_part_number": "M-001",
                "minimum_stock": 0,
            },
            actor_id=admin.id,
        )
        other_item = await CatalogService(unit_of_work).create_item(
            {
                "organization_id": other_organization.id,
                "sku": "ALIEN-001",
                "name": "Alien",
                "category_id": other_category.id,
                "unit_of_measure_id": other_unit.id,
                "item_type": "component",
                "minimum_stock": 0,
            },
            actor_id=admin.id,
        )
    return {
        "organization": organization,
        "other_organization": other_organization,
        "admin": admin,
        "viewer": viewer,
        "unit": unit,
        "item": item,
        "other_item": other_item,
        "other_unit": other_unit,
    }


async def create_spec_with_lines(db_session: AsyncSession) -> dict[str, object]:
    data = await create_bom_fixture(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        spec = await service.create_specification(
            {
                "organization_id": data["organization"].id,
                "code": "DZH-BOM-001",
                "name": "Специфікація БпАК «Джміль»",
                "product_item_id": data["item"].id,
                "specification_type": "product",
                "notes": "Для виробництва",
                "is_active": True,
            },
            actor_id=data["admin"].id,
        )
        version = await service.current_version(spec.id)
        inventory_line = await service.add_line(
            version.id,
            {
                "inventory_item_id": data["item"].id,
                "quantity": 2,
                "unit_of_measure_id": data["unit"].id,
            },
            actor_id=data["admin"].id,
        )
        manual_line = await service.add_line(
            version.id,
            {
                "display_name": "Кабель живлення ручний",
                "position_code": "MAN-001",
                "quantity": 1,
                "unit_of_measure_id": data["unit"].id,
                "notes": "Без створення номенклатури",
                "source_type": "manual",
            },
            actor_id=data["admin"].id,
        )
    data.update(
        {
            "spec": spec,
            "version": version,
            "inventory_line": inventory_line,
            "manual_line": manual_line,
        }
    )
    return data


async def test_create_specification_duplicate_code_and_lines(db_session: AsyncSession) -> None:
    data = await create_spec_with_lines(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        specs, total = await service.list_specifications(
            filters={"organization_id": data["organization"].id},
            page=PageRequest(page=1, page_size=50),
            sort_by="code",
            sort_direction=SortDirection.ASC,
        )
        lines = await service.list_lines(data["version"].id)
        with pytest.raises(ConflictError):
            await service.create_specification(
                {
                    "organization_id": data["organization"].id,
                    "code": "DZH-BOM-001",
                    "name": "Duplicate",
                    "specification_type": "product",
                    "is_active": True,
                },
                actor_id=data["admin"].id,
            )

    assert total == 1
    assert specs[0].name == "Специфікація БпАК «Джміль»"
    assert lines[0].display_name == "Двигун"
    assert lines[1].source_type == "manual"


async def test_line_display_name_does_not_rename_inventory_item(db_session: AsyncSession) -> None:
    data = await create_spec_with_lines(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        updated = await service.update_line(
            data["version"].id,
            data["inventory_line"].id,
            {"display_name": "Двигун лівого крила"},
            expected_version=data["inventory_line"].version,
            actor_id=data["admin"].id,
        )
        item = await CatalogService(unit_of_work).get_item(data["item"].id)
        audits = (
            await db_session.scalars(
                select(AuditLog).where(AuditLog.entity_id == data["inventory_line"].id)
            )
        ).all()

    assert updated.display_name == "Двигун лівого крила"
    assert item.name == "Двигун"
    assert any(log.before_data and log.before_data["display_name"] == "Двигун" for log in audits)


async def test_reorder_approve_immutability_and_new_version_snapshot(
    db_session: AsyncSession,
) -> None:
    data = await create_spec_with_lines(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        reordered = await service.reorder_lines(
            data["version"].id,
            [data["manual_line"].id, data["inventory_line"].id],
            actor_id=data["admin"].id,
        )
        approved = await service.approve_version(data["version"].id, actor_id=data["admin"].id)
        snapshot_before = await service.prepare_document(approved.id)
        with pytest.raises(ValidationError):
            await service.update_line(
                approved.id,
                data["inventory_line"].id,
                {"display_name": "Неможлива зміна"},
                expected_version=None,
                actor_id=data["admin"].id,
            )
        version2 = await service.create_version(
            data["spec"].id,
            {"source_version_id": approved.id, "change_reason": "Оновлення норми"},
            actor_id=data["admin"].id,
        )
        lines_v2 = await service.list_lines(version2.id)
        await service.update_line(
            version2.id,
            lines_v2[0].id,
            {"display_name": "Кабель живлення версія 2", "quantity": 3},
            expected_version=lines_v2[0].version,
            actor_id=data["admin"].id,
        )
        snapshot_after = await service.prepare_document(approved.id)

    assert [line.id for line in reordered] == [data["manual_line"].id, data["inventory_line"].id]
    assert snapshot_before == snapshot_after
    assert snapshot_after["lines"][0]["display_name"] == "Кабель живлення ручний"


async def test_approved_specification_header_and_attachments_are_immutable(
    db_session: AsyncSession,
) -> None:
    data = await create_spec_with_lines(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        approved = await service.approve_version(data["version"].id, actor_id=data["admin"].id)
        specification = await service.get_specification(data["spec"].id)
        with pytest.raises(ValidationError):
            await service.update_specification(
                specification.id,
                {"name": "Недозволена зміна"},
                expected_version=specification.version,
                actor_id=data["admin"].id,
            )
        with pytest.raises(ValidationError):
            await service.add_attachment(
                approved.id,
                {
                    "filename": "drawing.pdf",
                    "storage_key": "bom/drawing.pdf",
                    "mime_type": "application/pdf",
                    "file_size": 1024,
                },
                actor_id=data["admin"].id,
            )


async def test_archive_version_makes_current_specification_read_only(
    db_session: AsyncSession,
) -> None:
    data = await create_spec_with_lines(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        approved = await service.approve_version(data["version"].id, actor_id=data["admin"].id)
        snapshot_before = await service.prepare_document(approved.id)
        archived = await service.archive_version(approved.id, actor_id=data["admin"].id)
        specification = await service.get_specification(data["spec"].id)
        snapshot_after = await service.prepare_document(archived.id)
        with pytest.raises(ValidationError):
            await service.create_version(
                specification.id,
                {"source_version_id": archived.id},
                actor_id=data["admin"].id,
            )

    assert archived.status == "archived"
    assert specification.status == "archived"
    assert specification.is_active is False
    assert snapshot_before == snapshot_after


async def test_attachment_file_types_allow_cad_images_and_documents(
    db_session: AsyncSession,
) -> None:
    data = await create_spec_with_lines(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        dwg = await service.add_attachment(
            data["version"].id,
            {
                "filename": "frame.dwg",
                "storage_key": "bom/frame.dwg",
                "mime_type": "application/octet-stream",
                "file_size": 2048,
            },
            actor_id=data["admin"].id,
        )
        dxf = await service.add_attachment(
            data["version"].id,
            {
                "filename": "plate.dxf",
                "storage_key": "bom/plate.dxf",
                "mime_type": "application/dxf",
                "file_size": 2048,
            },
            actor_id=data["admin"].id,
        )
        docx = await service.add_attachment(
            data["version"].id,
            {
                "filename": "requirements.docx",
                "storage_key": "bom/requirements.docx",
                "mime_type": (
                    "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                ),
                "file_size": 2048,
            },
            actor_id=data["admin"].id,
        )

    assert dwg.filename == "frame.dwg"
    assert dxf.filename == "plate.dxf"
    assert docx.filename == "requirements.docx"


async def test_exports_single_specification_and_ukrainian_text(db_session: AsyncSession) -> None:
    data = await create_spec_with_lines(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        document = await service.prepare_document(data["version"].id)
        pdf = render_pdf(document)
        xlsx = render_xlsx(document)
        preview_html = render_preview_html(document, include_toolbar=False)
        filename = export_filename(document["specification"]["code"], 1, "pdf")

    assert pdf.startswith(b"%PDF")
    assert b"????" not in pdf
    assert b"/ToUnicode" in pdf
    assert b"/FontFile2" in pdf
    assert xlsx.startswith(b"PK")
    workbook = load_workbook(BytesIO(xlsx), data_only=True)
    sheet = workbook[WORKSHEET_NAME]
    assert sheet["A1"].value == "ТОВ «Укрфлайбуд»"
    assert sheet["A2"].value == "СПЕЦИФІКАЦІЯ ВИРОБУ"
    assert sheet["B5"].value == "DZH-BOM-001"
    assert sheet["C10"].value == "Найменування"
    assert "A1:G1" in {str(range_) for range_ in sheet.merged_cells.ranges}
    assert "A2:G2" in {str(range_) for range_ in sheet.merged_cells.ranges}
    assert "A3:G3" in {str(range_) for range_ in sheet.merged_cells.ranges}
    assert sheet.column_dimensions["C"].width >= 42
    assert sheet.column_dimensions["G"].width >= 36
    assert sheet.freeze_panes == "A11"
    assert sheet.auto_filter.ref == f"A10:G{sheet.max_row}"
    assert sheet.print_title_rows == "$10:$10"
    assert sheet.page_setup.orientation == sheet.ORIENTATION_LANDSCAPE
    assert sheet.print_area == f"'{WORKSHEET_NAME}'!$A$1:$G${sheet.max_row}"
    assert "bom-print-document" in preview_html
    assert "<thead>" in preview_html
    assert "../export/pdf" not in preview_html
    assert filename.startswith("SPEC_DZH-BOM-001_v1_")


async def test_preview_pdf_excel_api_authorization(db_session: AsyncSession) -> None:
    data = await create_spec_with_lines(db_session)
    app = create_app()

    async def override_unit_of_work() -> AsyncGenerator[UnitOfWorkStub]:
        async with service_context(db_session) as unit_of_work:
            yield unit_of_work

    app.dependency_overrides[get_unit_of_work] = override_unit_of_work
    admin_token = create_access_token(data["admin"].id, set())
    viewer_token = create_access_token(data["viewer"].id, set())
    headers = {"Authorization": f"Bearer {admin_token}"}
    forbidden_headers = {"Authorization": f"Bearer {viewer_token}"}
    version_id = data["version"].id

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        preview = await client.get(
            f"/api/v1/bom/versions/{version_id}/preview",
            headers=headers,
            params={"toolbar": False},
        )
        pdf = await client.get(f"/api/v1/bom/versions/{version_id}/export/pdf", headers=headers)
        xlsx = await client.get(f"/api/v1/bom/versions/{version_id}/export/xlsx", headers=headers)
        forbidden_preview = await client.get(
            f"/api/v1/bom/versions/{version_id}/preview", headers=forbidden_headers
        )
        forbidden_pdf = await client.get(
            f"/api/v1/bom/versions/{version_id}/export/pdf", headers=forbidden_headers
        )
        forbidden_xlsx = await client.get(
            f"/api/v1/bom/versions/{version_id}/export/xlsx", headers=forbidden_headers
        )

    assert preview.status_code == 200
    assert "text/html" in preview.headers["content-type"]
    assert "DZH-BOM-001" in preview.text
    assert "../export/pdf" not in preview.text
    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF")
    assert xlsx.status_code == 200
    assert xlsx.content.startswith(b"PK")
    assert forbidden_preview.status_code == 403
    assert forbidden_pdf.status_code == 403
    assert forbidden_xlsx.status_code == 403


async def test_import_validation_no_partial_import_on_error(db_session: AsyncSession) -> None:
    data = await create_spec_with_lines(db_session)
    content = render_import_template()
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    sheet.append(["NEW-001", "Нова позиція", 5, "шт", "UFB", "P-1", "", "", ""])
    sheet.append(["BAD-001", "", 0, "шт", "", "", "", "", ""])
    output = BytesIO()
    workbook.save(output)

    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        result = await service.import_lines(
            data["version"].id, output.getvalue(), actor_id=data["admin"].id
        )
        lines = await service.list_lines(data["version"].id)

    assert result["valid"] is False
    assert len(lines) == 2


async def test_organization_isolation_and_permissions(db_session: AsyncSession) -> None:
    data = await create_bom_fixture(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        spec = await service.create_specification(
            {
                "organization_id": data["organization"].id,
                "code": "ORG-BOM-1",
                "name": "Організаційна ізоляція",
                "specification_type": "product",
                "is_active": True,
            },
            actor_id=data["admin"].id,
        )
        version = await service.current_version(spec.id)
        with pytest.raises(ValidationError):
            await service.add_line(
                version.id,
                {
                    "inventory_item_id": data["other_item"].id,
                    "quantity": 1,
                    "unit_of_measure_id": data["unit"].id,
                },
                actor_id=data["admin"].id,
            )
        await IdentityService(unit_of_work).seed_system_access()
        with pytest.raises(PermissionDeniedError):
            await require_permission_for_user(unit_of_work, data["viewer"].id, "bom.export")
        role = await db_session.scalar(select(RoleModel).where(RoleModel.code == "bom_viewer"))
        bom_read = await db_session.scalar(
            select(PermissionModel).where(PermissionModel.code == "bom.read")
        )
        assert role is not None
        assert bom_read is not None
        role_permission = await db_session.scalar(
            select(RolePermissionModel).where(
                RolePermissionModel.role_id == role.id,
                RolePermissionModel.permission_id == bom_read.id,
            )
        )
        assert role_permission is not None


async def test_bootstrap_admin_token_contains_canonical_bom_permissions(
    db_session: AsyncSession, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(settings, "bootstrap_admin_email", "admin@example.com")
    monkeypatch.setattr(settings, "bootstrap_admin_name", "Administrator")
    monkeypatch.setattr(settings, "bootstrap_admin_password", "ChangeMe12345")
    async with service_context(db_session) as unit_of_work:
        await IdentityService(unit_of_work).bootstrap_admin()
        access, _, _, _ = await AuthService(unit_of_work).login(
            email=settings.bootstrap_admin_email,
            password=settings.bootstrap_admin_password,
            ip_address="127.0.0.1",
            user_agent="pytest",
        )

    permissions = set(decode_access_token(access)["permissions"])
    assert {
        "bom.read",
        "bom.create",
        "bom.edit",
        "bom.approve",
        "bom.export",
        "bom.import",
        "bom.attachments",
    }.issubset(permissions)


async def test_add_new_position_and_template_download(db_session: AsyncSession) -> None:
    data = await create_spec_with_lines(db_session)
    async with service_context(db_session) as unit_of_work:
        service = BomService(unit_of_work)
        added = await service.add_line(
            data["version"].id,
            {
                "display_name": "Додана позиція",
                "quantity": 4,
                "unit_of_measure_id": data["unit"].id,
            },
            actor_id=data["admin"].id,
        )
    assert added.display_name == "Додана позиція"
    assert render_import_template().startswith(b"PK")
    assert uuid4() != added.id
